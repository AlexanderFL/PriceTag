# Python 3.6.4

from openpyxl import load_workbook
debug = True  # Setting debug to true will disable user input


def is_debug():
    return debug


class Labels:
    hillumidi = 16
    stormidi = 2
    a6b = 4


nr_of_labels = 16
if not is_debug():
    nr_of_labels = int(input(": "))

label_type = Labels.stormidi

code_filename = 'vorunr.txt'
excel_filename = 'test1.xlsm'

wb = None

try:
    wb = load_workbook(excel_filename, keep_vba=True)
except PermissionError:
    print("Access was denied, is the file open? If so, close the file and try again.")
    exit()
except FileNotFoundError:
    print("File '" + excel_filename + "' was not found.")
    exit()

ws = wb.active

file = open(code_filename, "r")
codes = []
for line in file:
    codes.append(line)
print(codes)

label_count = 0
label_sub_count = 0
while label_count != len(codes):
    if label_sub_count % label_type == 0 and label_count is not 0:
        label_sub_count = 0
        print("printing")
    col_string = 'A' + str(2+label_sub_count)
    print(col_string)
    ws[col_string] = codes[label_count]

    label_sub_count += 1
    label_count += 1
print("printing")

wb.template = False
wb.save('test1.xlsm')
